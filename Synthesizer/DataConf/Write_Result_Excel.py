# -*- coding: utf-8 -*-
'''
Created on 2019年9月3日

@author: likecan
'''

import openpyxl
from xlutils.copy import copy
from DataConf import Total_LET, Part_Result
from MySQLdb._mysql import result




class write_result_to_excel():
    
    all_titles = {'GSM':-1,'LTE':-1,'CDMA':-1,'WCDMA':-1,'TDSCDMA':-1}
    sheet_arry = []
    
    
    def __init__(self,excel_path='D:\\',excel_name='综测仪结果图表.xlsx'):
        self.result_data = result_data
        if excel_path[-1] == '//':
            excel_path = excel_path+'//'
        self.excel_whole_path = excel_path+excel_name
        self.result_excel = openpyxl.load_workbook(self.excel_whole_path)
        self.sheet_arry = self.result_excel.sheetnames
        print(self.sheet_arry)
        self.total_sheet = self.result_excel['total']
        #self.network_type = result_data[0]['NetWork']
        #sheet_name +='_'+self.network_type.upper()
        #self.result_excel_sheet = self.result_excel[sheet_name]
            
    
    def sheetnames_operation(self):
        result_data_arry = []
        if not self.sheet_arry:
            print('no sheet in excel!')
        else:
            for sheet_name in self.sheet_arry:
                if '-' in sheet_name:
                    result_excel_sheet = self.result_excel[sheet_name]
                    for rows in range(3,result_excel_sheet.max_row+1):
                        result_data = {}
                        network_type = result_excel_sheet.cell(rows,Part_Result.get_network_index()+1).value
                        if not network_type:
                            continue
                        result_data['NetWork'] = network_type
                        result_data['Test_Item'] = result_excel_sheet.cell(rows,Part_Result.get_band_index()+1).value
                        result_data['Max_Delta_Sp0ec'] = result_excel_sheet.cell(rows,Part_Result.get_whole_ave_current_index()+1).value
                        result_data['Max_Delta_Curr'] = result_excel_sheet.cell(rows,Part_Result.get_whole_integrtion_current_index()+1).value
                        result_data['Min_Spec'] = result_excel_sheet.cell(rows,Part_Result.get_whole_pa_current_index()+1).value
                        result_data['Min_Current'] = result_excel_sheet.cell(rows,Part_Result.get_whole_pa_current_index()+1).value
                        result_data['Max_ResultJudge'] = self.__result_judge(result_excel_sheet.cell(rows,Part_Result.get_plc_or_judgement()))
                        result_data['Min_ResultJudge'] = self.__result_judge(result_excel_sheet.cell(rows,Part_Result.get_plc_or_judgement()))
                        if result_data['NetWork'] == 'LTE':
                            result_data['APT_10dBm_Delta_Spec'] = result_excel_sheet.cell(rows,Part_Result.get_whole_ave_current_index()+1).value
                            result_data['APT_10dBm_Delta_Curr'] = result_excel_sheet.cell(rows,Part_Result.get_whole_integrtion_current_index()+1).value
                            result_data['CL_3dBm_Delta_Spec'] = result_excel_sheet.cell(rows,Part_Result.get_whole_pa_current_index()+1).value
                            result_data['CL_3dBm_Delta_Curr'] = result_excel_sheet.cell(rows,Part_Result.get_whole_pa_current_index()+1).value
                            result_data['APT_ResultJudge'] = self.__result_judge(result_excel_sheet.cell(rows,Part_Result.get_plc_or_judgement()))
                            result_data['CL_ResultJudge'] = self.__result_judge(result_excel_sheet.cell(rows,Part_Result.get_plc_or_judgement()))
                        result_data_arry.append(result_data)
        print(result_data_arry)
        return result_data_arry
    
    def write_total_result_by_network_type(self,result_data):
        try:
            for index in range (len(result_data)):#从传入的数据当中取值
                insert_row = self.check_insert_row(result_data[0]['NetWork'])
                print('NetWork  '+result_data[0]['NetWork']+': insert_row:'+str(insert_row))
                self.total_sheet.insert_rows(insert_row)
                insert_colum = 1
                result_data_content = result_data[index]
                for key in result_data_content:#获取具体的数据
                    if key == 'NetWork':
                        continue
                    data_content = result_data_content[key]
                    self.total_sheet.cell(insert_row,insert_colum,data_content)#插入数据
                    insert_colum += 1
                    #self.__titles_add(index, self.all_titles[self.result_data[index]['NetWork']])
        except Exception as e:
            print(e)
        finally:
            self.result_excel.save(self.excel_whole_path)
            
            
            
            
    def __result_judge(self,judgement):
        return 'Pass'
    

    def check_insert_row(self,network_type):
        data_insert_row = 0
        for rows in range(1,self.total_sheet.max_row+1):
            first_col_data = self.total_sheet.cell(rows,1).value
            if network_type == first_col_data:
                data_insert_row = rows+3
                break
        return data_insert_row        
        



    
    


result_data = [{'NetWork':'LTE','Test_Item':'B12','Max_Delta_Sp0ec':'5','Max_Delta_Curr':'200','Max_ResultJudge':'Fail','APT_10dBm_Delta_Spec':7,'APT_10dBm_Delta_Curr':300,'APT_ResultJudge':'success',
               'CL_3dBm_Delta_Spec':8,'CL_3dBm_Delta_Curr':1201,'CL_ResultJudge':'Fail','Min_Spec':1,'Min_Current':121,'Min_ResultJudge':'success'},{'NetWork':'LTE','Test_Item':'B3','Max_Delta_Spec':'5','Max_Delta_Curr':'200','Max_ResultJudge':'Fail','APT_10dBm_Delta_Spec':7,'APT_10dBm_Delta_Curr':300,'APT_ResultJudge':'success',
               'CL_3dBm_Delta_Spec':8,'CL_3dBm_Delta_Curr':1201,'CL_ResultJudge':'Fail','Min_Spec':1,'Min_Current':121,'Min_ResultJudge':'success'},{'NetWork':'LTE','Test_Item':'B3','Max_Delta_Spec':'5','Max_Delta_Curr':'200','Max_ResultJudge':'Fail','APT_10dBm_Delta_Spec':7,'APT_10dBm_Delta_Curr':300,'APT_ResultJudge':'success',
               'CL_3dBm_Delta_Spec':8,'CL_3dBm_Delta_Curr':1201,'CL_ResultJudge':'Fail','Min_Spec':1,'Min_Current':121,'Min_ResultJudge':'success'},{'NetWork':'GSM','Test_Item':'1800','Max_Delta_Spec':'5','Max_Delta_Curr':'200','Max_ResultJudge':'Fail','Min_Spec':1,'Min_Current':121,'Min_ResultJudge':'success'},
               {'NetWork':'TDSCDMA','Test_Item':'B3','Max_Delta_Spec':'5','Max_Delta_Curr':'200','Max_ResultJudge':'Fail','Min_Spec':1,'Min_Current':121,'Min_ResultJudge':'success'}]
wrte = write_result_to_excel()
wrte.write_total_result_by_network_type(result_data)
print('END')
